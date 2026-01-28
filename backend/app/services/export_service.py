import csv
import io
from typing import List, Optional
from uuid import UUID
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.search import Search, SearchResult
from app.models.influencer import Influencer
from app.core.exceptions import ExportError


class ExportService:
    """Service for exporting search results to various formats."""

    EXPORT_COLUMNS = [
        ("Rank", "rank_position"),
        ("Username", "username"),
        ("Display Name", "display_name"),
        ("Relevance Score", "relevance_score"),
        ("Credibility %", "credibility_score"),
        ("Engagement Rate %", "engagement_rate"),
        ("Spain Audience %", "spain_audience_pct"),
        ("6M Growth %", "growth_rate"),
        ("Followers", "follower_count"),
        ("Male Audience %", "male_pct"),
        ("Female Audience %", "female_pct"),
        ("Avg Likes", "avg_likes"),
        ("Avg Comments", "avg_comments"),
        ("Interests", "interests"),
        ("Profile URL", "profile_url"),  # Derived from username
    ]

    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_to_csv(self, search_id: UUID) -> str:
        """Export search results to CSV format."""
        results = await self._get_results_with_data(search_id)

        if not results:
            raise ExportError(f"No results found for search {search_id}")

        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([col[0] for col in self.EXPORT_COLUMNS])

        # Data rows
        for result in results:
            writer.writerow([
                self._format_value(result, col[1])
                for col in self.EXPORT_COLUMNS
            ])

        return output.getvalue()

    async def export_to_excel(self, search_id: UUID) -> bytes:
        """Export search results to Excel format."""
        results = await self._get_results_with_data(search_id)
        search = await self._get_search(search_id)

        if not results:
            raise ExportError(f"No results found for search {search_id}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Influencer Results"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add search info header
        ws.merge_cells('A1:N1')
        ws['A1'] = f"Search: {search.raw_query}" if search else "Search Results"
        ws['A1'].font = Font(bold=True, size=14)

        ws.merge_cells('A2:N2')
        ws['A2'] = f"Exported: {search.executed_at.strftime('%Y-%m-%d %H:%M') if search else 'N/A'}"

        # Header row (row 4)
        for col_idx, (col_name, _) in enumerate(self.EXPORT_COLUMNS, 1):
            cell = ws.cell(row=4, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, result in enumerate(results, 5):
            for col_idx, (_, field) in enumerate(self.EXPORT_COLUMNS, 1):
                value = self._format_value(result, field)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Format numbers
                if field in ['relevance_score', 'credibility_score', 'engagement_rate',
                            'spain_audience_pct', 'growth_rate', 'male_pct', 'female_pct']:
                    cell.number_format = '0.00'
                elif field in ['follower_count', 'avg_likes', 'avg_comments']:
                    cell.number_format = '#,##0'

        # Adjust column widths
        column_widths = [8, 20, 25, 15, 15, 18, 18, 12, 15, 15, 15, 12, 14, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Freeze header row
        ws.freeze_panes = 'A5'

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def _get_results_with_data(self, search_id: UUID) -> List[dict]:
        """Get search results with full influencer data."""
        # Get search results with influencer data
        query = (
            select(SearchResult, Influencer)
            .join(Influencer, SearchResult.influencer_id == Influencer.id)
            .where(SearchResult.search_id == search_id)
            .order_by(SearchResult.rank_position)
        )

        result = await self.db.execute(query)
        rows = result.all()

        results = []
        for search_result, influencer in rows:
            # Extract metrics
            genders = influencer.audience_genders or {}
            geography = influencer.audience_geography or {}

            results.append({
                'rank_position': search_result.rank_position,
                'username': influencer.username,
                'display_name': influencer.display_name,
                'relevance_score': search_result.relevance_score,
                'credibility_score': influencer.credibility_score,
                'engagement_rate': (influencer.engagement_rate or 0) * 100,  # Convert to percentage
                'spain_audience_pct': geography.get('ES', 0),
                'growth_rate': (influencer.follower_growth_rate_6m or 0) * 100,  # Convert to percentage
                'follower_count': influencer.follower_count,
                'male_pct': genders.get('male', genders.get('Male', 0)),
                'female_pct': genders.get('female', genders.get('Female', 0)),
                'avg_likes': influencer.avg_likes,
                'avg_comments': influencer.avg_comments,
                'interests': ', '.join(influencer.interests or []),
                'profile_url': f"https://instagram.com/{influencer.username}",  # Derived from username
            })

        return results

    async def _get_search(self, search_id: UUID) -> Optional[Search]:
        """Get search record."""
        query = select(Search).where(Search.id == search_id)
        result = await self.db.execute(query)
        return result.scalar_one_or_none()

    def _format_value(self, result: dict, field: str):
        """Format a value for export."""
        value = result.get(field)

        if value is None:
            return ""

        if field == 'relevance_score':
            return round(value * 100, 2)  # Convert to percentage

        if field in ['credibility_score', 'spain_audience_pct', 'male_pct', 'female_pct']:
            return round(value, 2) if value else 0

        if field in ['engagement_rate', 'growth_rate']:
            return round(value, 2) if value else 0

        if field in ['follower_count', 'avg_likes', 'avg_comments', 'rank_position']:
            return int(value) if value else 0

        return value
